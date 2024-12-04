#!/usr/bin/env python3
from cobra import Model, Reaction, Metabolite
from openpyxl import load_workbook
import json, sys, re

# Load reactions from the ModelSEED database (reference)
f = open("../files/ModelSEED_db_reactions.json", "r")
modelSEED_reactions = json.load(f)

# Load compounds from the ModelSEED database (reference)
g = open("../files/ModelSEED_db_compounds.json", "r")
modelSEED_compounds = json.load(g)

# Function to clean up a list
def remove_blanks_and_pluses(a_list):
    for index, entry in enumerate(a_list):
        repl = entry
        if entry.count("+") > 1:
            repl = entry.split("+")[0]
        repl = repl.strip()
        if "] +" in repl:
            repl = repl[:-2]
        a_list[index] = repl
    return a_list

# Function to handle manual reactions - might to be omitted if not functional
def parse_manual_reaction(tmp, part):

    if tmp[0] != "(":
        if part == "r":
            stoichiometry = -1
        else:
            stoichiometry = 1
    else:
        end = tmp.index(")")
        if part == "r":
            stoichiometry = -int(tmp[1:end])
        else:
            stoichiometry = tmp[1:end]

    z = re.compile("\(\d\)").split(tmp)
    module = list(filter(None, z))
    compound = module[0].strip()
    if compound == "H+[1]":
        compound = "H+"
    return ([stoichiometry, compound])


# Main function to parse a sheet with reactions
def build_toy_model(sheet_with_reactions, model):

    # Reading sheet per row
    reactions_with_rxn_id = []
    manual_reactions         = []
    set_of_metabolites       = set()

    for row in sheet_with_reactions.iter_rows(min_row=1, max_col=5, values_only=True):
        reaction_name = row[2]
        if row[1] != None:
            for entry in modelSEED_reactions:

                if entry["name"] == reaction_name:
                    entry["upper_bound"] = 1000
                    entry["lower_bound"] = -1000

                    # Add compounds
                    switch = False
                    comp   = "c"
                    to_be_exchanged = []
                    for participant in entry["stoichiometry"].split(";"):
                        metabolite = participant.split(":")[1]
                        for term in modelSEED_compounds:
                            mt_id = term["id"] + "_" + comp
                            if term["id"] == metabolite:
                                if row[4] != None:
                                    switch = True

                                model_metabolite = Metabolite(mt_id, \
                                                            name        = term["name"], \
                                                            formula     = term["formula"], \
                                                            compartment = comp)
                                model.add_metabolites([model_metabolite])
                                set_of_metabolites.add(mt_id)

                                if switch:
                                    comp = "e"
                                    mt_id = term["id"] + "_" + comp
                                    model_metabolite = Metabolite(mt_id, \
                                                                name        = term["name"], \
                                                                formula     = term["formula"], \
                                                                compartment = comp)
                                    model.add_metabolites([model_metabolite])
                                    set_of_metabolites.add(mt_id)
                                    to_be_exchanged.append(mt_id)
                    # Add reaction
                    if row[4] != None:
                        model.add_boundary(model.metabolites.get_by_id(to_be_exchanged[0]), type = "exchange")
                    else:
                        reaction = Reaction(entry["id"])
                        model.add_reactions([reaction])
                        reaction.name = entry["name"]
                        reaction.lower_bound = entry["lower_bound"]
                        reaction.upper_bound = entry["upper_bound"]

                        for participant in entry["stoichiometry"].split(";"):
                            metabolite    = participant.split(":")[1] + "_c"
                            stoichiometry = participant.split(":")[0]
                            reaction.add_metabolites({metabolite : int(stoichiometry)})

                    reactions_with_rxn_id.append(entry)
        else:
            manual_reactions.append((row[3], row[4], row[0]))

    # Parse manual reactions
    for case in manual_reactions:
        check = True
        reaction = case[0]
        exchange = case[1]
        react_id = case[2]

        modules = []

        react_part    = re.split('<=>|<=|=>', reaction)[0].strip()
        product_part  = re.split('<=>|<=|=>', reaction)[1].strip()

        tmp = react_part.split(" + ")

        for i in tmp:
            module = parse_manual_reaction(i, "r")
            modules.append(module)

        tmp = product_part.split(" + ")
        for i in tmp:
            module = parse_manual_reaction(i, "p")
            modules.append(module)

        # Add compounds to the model
        for index, module in enumerate(modules):
            if check == False:
                break
            compound         = module[1]
            compound_in_SEED = False

            for modelSEED_compound in modelSEED_compounds:

                if modelSEED_compound["name"] == compound:
                    # Check whether this is an exchange reaction and you need only a sole metabolite
                    if react_part == product_part and "EX_" in row[0]:

                        mt_id            = compounds_to_ids[compound] + "_e"
                        model_metabolite = Metabolite(mt_id, \
                                                    name        = mt_id,\
                                                    compartment = comp)
                        model.add_metabolites([model_metabolite])

                        react = Reaction('EX_' + mt_id)
                        model.add_reactions([react])
                        react.name = react_id
                        react.lower_bound = -1000
                        react.upper_bound = 1000
                        react.add_metabolites({mt_id:-1})
                        check = False
                        compound_in_SEED = True
                        break

                    else:
                        compound_in_SEED = True
                        comp             = "c"
                        sbml_compound    = compounds_to_ids[compound]

                        mt_id = sbml_compound + "_" + comp

                        model_metabolite = Metabolite(mt_id, \
                                                    name        = mt_id,\
                                                    compartment = comp)
                        model.add_metabolites([model_metabolite])
                        modules[index][1] = mt_id

                        if exchange != None:
                            comp = "e"
                            mt_id = sbml_compound + "_" + comp
                            model_metabolite = Metabolite(mt_id, \
                                                        name        = mt_id,\
                                                        compartment = comp)
                            model.add_metabolites([model_metabolite])
                        break

            if compound_in_SEED == False:
                print(compound, " not in ModelSEED")

        if check == False:
            continue

        # Add reaction to the model
        reaction = Reaction(react_id)
        model.add_reactions([reaction])
        reaction.name = react_id
        reaction.lower_bound = -1000
        reaction.upper_bound = 1000

        for participant in modules:
            metabolite    = participant[1]
            stoichiometry = participant[0]
            reaction.add_metabolites({metabolite : int(stoichiometry)})

    #model.objective = 'biomass'
    return model


# --------------------
# Part A: ModelSEED ids - names dictionary
# --------------------

# Parse ModelSEED database reactions to map compounds with their corresponding ids
compounds_to_ids = {}
for reaction in modelSEED_reactions:

    definition = reaction["definition"]
    equation   = reaction["equation"]

    definition = re.split('<=>|<=|=>', definition)

    def_comp = remove_blanks_and_pluses(list(filter(None, re.compile("\(\d\)").split(definition[0]))))
    def_prod = remove_blanks_and_pluses(list(filter(None, re.compile("\(\d\)").split(definition[1]))))

    compounds_names = []
    for x in list(filter(None, def_comp + def_prod)):
        compounds_names.append(x.split("[")[0])

    compounds_ids   = []
    for entry in equation.split(" "):
        if "cpd" in entry:
            compounds_ids.append(entry.split("[")[0])

    for k, v in zip(compounds_names, compounds_ids):
        compounds_to_ids[k] = v

# with open('../files/comp_names_to_modelSEED_ids.json', 'w') as outfile:
#     json.dump(compounds_to_ids, outfile)



# --------------------
# Part B1: Glucose fermenter (bt)
# --------------------


#Init model
init_model = Model('toy_bt')

# Read toy model's reactions
# my_reactions_excel = load_workbook(filename = 'metabolicReactions2.xlsx')
my_reactions_excel = load_workbook(filename = '../files/speciesSpecificReactions.xlsx')
# bt_reactions_sheet = my_reactions_excel["BT_metabolicReactions"]
bt_reactions_sheet = my_reactions_excel["bt"]

# Build model
bt_model = build_toy_model(bt_reactions_sheet, init_model)

#Validate and write model in a .xml file
from pprint import pprint
from cobra.io import write_sbml_model, validate_sbml_model
with open('../files/models/bt2_toy_model.xml', "w") as f_sbml:
    write_sbml_model(bt_model, filename = f_sbml.name)
    report = validate_sbml_model(filename = f_sbml.name)

pprint(report)
sys.exit(0)

# --------------------
# Part B2: butyrate producer (ri)
# --------------------

# # Init model
# init_model = Model('toy_ri')

# # Read toy model's reactions
# my_reactions_excel = load_workbook(filename = 'metabolicReactions.xlsx')
# ri_reactions_sheet = my_reactions_excel["ri"]

# # Build model
# ri_model = build_toy_model(ri_reactions_sheet, init_model)
# cobra.io.write_sbml_model(ri_model, 'ri_toy_model2.xml')
# # Validate and write model in a .xml file
# with open('../files/models/ri_toy_model.xml', "w") as f_sbml:
#     write_sbml_model(ri_model, filename = f_sbml.name)
#     report = validate_sbml_model(filename = f_sbml.name)
# pprint(report)


# --------------------
# Part B3: acetogen (bh)
# --------------------

# Init model
# init_model = Model('toy_bh')

# # Read toy model's reactions
# bh_reactions_sheet = my_reactions_excel["bh"]

# # Build model
# bh_model = build_toy_model(bh_reactions_sheet, init_model)

# Validate and write model in a .xml file
# with open('../files/models/bh_toy_model.xml', "w") as f_sbml:
#     write_sbml_model(bh_model, filename = f_sbml.name)
#     report = validate_sbml_model(filename = f_sbml.name)

# pprint(report)

