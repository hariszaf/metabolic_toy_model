#!/usr/bin/env python3
from cobra import Model, Reaction, Metabolite
from openpyxl import load_workbook
import json, sys, re

# Load reactions from the ModelSEED database (reference)
f = open("ModelSEED_db_reactions.json", "r")
modelSEED_reactions = json.load(f)

# Load compounds from the ModelSEED database (reference)
g = open("ModelSEED_db_compounds.json", "r")
modelSEED_compounds = json.load(g)

# Function to clean up a list 
def remove_blanks_and_pluses(a_list):
    for index, entry in enumerate(a_list):
        repl = entry
        if entry.count("+") > 1:
            repl = entry.split("+")[0]
        repl = repl.strip()
        if "] +" in repl:
            repl = repl[:-2]
        a_list[index] = repl
    return a_list


def parse_manual_reaction(tmp, part):

    if tmp[0] != "(":
        if part == "r":
            stoichiometry = -1
        else:
            stoichiometry = 1
    else:
        end = tmp.index(")")
        if part == "r":
            stoichiometry = -int(tmp[1:end])
        else:
            stoichiometry = tmp[1:end]

    z = re.compile("\(\d\)").split(tmp)
    module = list(filter(None, z))
    compound = module[0].strip()
    if compound == "H+[1]":
        compound = "H+"
    return (stoichiometry, compound)



# --------------------
# Part A: ModelSEED ids - names dictionary 
# --------------------

# Parse ModelSEED database reactions to map compounds with their corresponding ids 
compounds_to_ids = {}
for reaction in modelSEED_reactions:

    definition = reaction["definition"]
    equation   = reaction["equation"]

    definition = re.split('<=>|<=|=>', definition)

    def_comp = remove_blanks_and_pluses(list(filter(None, re.compile("\(\d\)").split(definition[0]))))
    def_prod = remove_blanks_and_pluses(list(filter(None, re.compile("\(\d\)").split(definition[1]))))

    compounds_names = []
    for x in list(filter(None, def_comp + def_prod)):
        compounds_names.append(x.split("[")[0])

    compounds_ids   = []
    for entry in equation.split(" "):
        if "cpd" in entry:
            compounds_ids.append(entry.split("[")[0])

    for k, v in zip(compounds_names, compounds_ids):
        compounds_to_ids[k] = v

with open('comp_names_to_modelSEED_ids.json', 'w') as outfile:
    json.dump(compounds_to_ids, outfile)



# --------------------
# Part B
# --------------------


# Init model 
model = Model('toy_bt')

# Set cases of extracellular 


# Read toy model's reactions 
my_reactions_excel = load_workbook(filename = 'metabolicReactions.xlsx')
bt_reactions_sheet = my_reactions_excel["bt"]

# Reading sheet per row
bt_reactions_with_rxn_id = []
manual_reactions         = []
set_of_metabolites       = set()

for row in bt_reactions_sheet.iter_rows(min_row=1, max_col=5, values_only=True):

    reaction_name = row[2]
    if row[1] != None:
        for entry in modelSEED_reactions:
            """
            example of an entry:
            {'abbreviation': 'GLCt2', 'abstract_reaction': None, 
            'aliases': ['BiGG: GLCt2; GLCt2pp', 'MetaCyc: RXN0-7077', 'iAF1260: GLCt2pp', 'iAG612: rbs_312; rbs_313', 'iAO358: rll_538', 'Name: D-glucose transport in via proton symport; D-glucose:proton symport; D-glucosetransportinviaprotonsymport'], 
            'code': '(1) cpd00027[1] <=> (1) cpd00027[0]', 
            'compound_ids': 'cpd00027;cpd00067', 
            'definition': '(1) D-Glucose[1] + (1) H+[1] <=> (1) D-Glucose[0] + (1) H+[0]', 
            'deltag': 0.0, 'deltagerr': 2.18, 'direction': '=', 'ec_numbers': None, 
            'equation': '(1) cpd00027[1] + (1) cpd00067[1] <=> (1) cpd00027[0] + (1) cpd00067[0]', 
            'id': 'rxn05573', 'is_obsolete': 0, 'is_transport': 1, 'linked_reaction': 'rxn08616', 'name': 'D-glucose transport in via proton symport', 
            'notes': ['GCC', 'EQC'], 'pathways': ['MetaCyc: Alcohol-Biosynthesis (Fermentation to Alcohols); Energy-Metabolism (Generation of Precursor Metabolite and Energy); Fermentation (); PWY-7385 (1,3-propanediol biosynthesis (engineered))'], 
            'reversibility': '=', 'source': 'Primary Database', 'status': 'OK', 
            'stoichiometry': '-1:cpd00027:1:0:"D-Glucose";-1:cpd00067:1:0:"H+";1:cpd00027:0:0:"D-Glucose";1:cpd00067:0:0:"H+"'}
            """
            if entry["name"] == reaction_name:
                entry["upper_bound"] = 1000
                entry["lower_bound"] = -1000

                # Add compounds
                switch = False
                comp   = "c"
                to_be_exchanged = []
                for participant in entry["stoichiometry"].split(";"):
                    metabolite = participant.split(":")[1]
                    for term in modelSEED_compounds:
                        mt_id = term["id"] + "_" + comp
                        if term["id"] == metabolite:
                            if row[4] != None:
                                switch = True 

                            model_metabolite = Metabolite(mt_id, \
                                                        name        = term["name"], \
                                                        formula     = term["formula"], \
                                                        compartment = comp)
                            model.add_metabolites([model_metabolite])
                            set_of_metabolites.add(mt_id)

                            if switch: 
                                comp = "e"
                                mt_id = term["id"] + "_" + comp
                                model_metabolite = Metabolite(mt_id, \
                                                            name        = term["name"], \
                                                            formula     = term["formula"], \
                                                            compartment = comp)
                                model.add_metabolites([model_metabolite])
                                set_of_metabolites.add(mt_id)
                                to_be_exchanged.append(mt_id)

                # Add reaction
                if row[4] != None:
                    model.add_boundary(model.metabolites.get_by_id(to_be_exchanged[0]), type = "exchange")

                else:
                    reaction = Reaction(entry["id"])
                    model.add_reactions([reaction])
                    reaction.name = entry["name"]
                    reaction.lower_bound = entry["lower_bound"]
                    reaction.upper_bound = entry["upper_bound"]

                    for participant in entry["stoichiometry"].split(";"):
                        metabolite    = participant.split(":")[1] + "_c"
                        stoichiometry = participant.split(":")[0]
                        reaction.add_metabolites({metabolite : int(stoichiometry)})


                bt_reactions_with_rxn_id.append(entry)
    else:
        manual_reactions.append((row[3], row[4], row[0]))



# Parse manual reactions
manual_compounds = []
counter = 0
for case in manual_reactions:
    reaction = case[0]
    exchange = case[1]
    react_id = case[2]

    modules = []

    react_part    = re.split('<=>|<=|=>', reaction)[0]
    product_part  = re.split('<=>|<=|=>', reaction)[1]

    tmp = react_part.split(" + ")

    for i in tmp:
        module = parse_manual_reaction(i, "r")
        modules.append(module)

    tmp = product_part.split(" + ")
    for i in tmp:
        module = parse_manual_reaction(i, "p")
        modules.append(module)

    # Add compounds to the model 
    for module in modules:
        compound = module[1]
        counter  += 1
        compound_in_SEED = False
        for modelSEED_compound in modelSEED_compounds: 
            if modelSEED_compound["name"] == compound: 
                compound_in_SEED = True
                comp = "c"
                if exchange != None:
                    if compound in exchange:
                        comp = "e"
                model_metabolite = Metabolite(compound, name = compound,\
                                              compartment = comp)
                model.add_metabolites([model_metabolite])
                break
        if compound_in_SEED == False:
            print(compound, " not in ModelSEED")

    # Add reaction to the model
    reaction = Reaction(react_id)
    model.add_reactions([reaction])
    reaction.name = react_id
    reaction.lower_bound = -1000
    reaction.upper_bound = 1000

    for participant in modules:
        metabolite    = participant[1]
        stoichiometry = participant[0]
        reaction.add_metabolites({metabolite : int(stoichiometry)})


for reaction in model.reactions:
    print(reaction)

model.objective = 'biomass'


# ---------------------
# Model validation & .xml writing
# ---------------------
from pprint import pprint
from cobra.io import write_sbml_model, validate_sbml_model
with open('bt_toy_model.xml', "w") as f_sbml:
    write_sbml_model(model, filename = f_sbml.name)
    report = validate_sbml_model(filename = f_sbml.name)

pprint(report)





